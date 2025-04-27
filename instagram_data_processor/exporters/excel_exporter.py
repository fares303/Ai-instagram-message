"""
Excel exporter for Instagram conversation data.
"""

import os
import logging
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import instagram_data_processor.utils as utils

logger = logging.getLogger(__name__)

class ExcelExporter:
    """
    Export conversation data to an Excel file.
    """

    def __init__(self, output_dir):
        """
        Initialize the Excel exporter.

        Args:
            output_dir (str): Directory to save the output file
        """
        self.output_dir = output_dir
        logger.info(f"Initialized Excel exporter with output directory: {output_dir}")

    def export(self, messages, target_user, my_name, stats=None):
        """
        Export messages to an Excel file.

        Args:
            messages (list): List of processed messages
            target_user (str): Name of the target user
            my_name (str): Your name
            stats (dict, optional): Statistics to include

        Returns:
            str: Path to the exported file
        """
        if not messages:
            logger.warning("No messages to export")
            return None

        # Create filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"conversation_with_{target_user}_{timestamp}.xlsx"
        filepath = os.path.join(self.output_dir, filename)

        try:
            # Create a Pandas Excel writer
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # Create the conversation sheet
                self._create_conversation_sheet(writer, messages, target_user, my_name)

                # Create the statistics sheet if stats are provided
                if stats:
                    self._create_statistics_sheet(writer, stats, target_user, my_name)

                # Create the media sheet
                self._create_media_sheet(writer, messages)

                # Apply styling to all sheets
                self._apply_styling(writer)

            logger.info(f"Exported conversation to Excel file: {filepath}")
            return filepath

        except Exception as e:
            logger.error(f"Error exporting to Excel file: {str(e)}")
            return None

    def _create_conversation_sheet(self, writer, messages, target_user, my_name):
        """
        Create the conversation sheet in the Excel file.

        Args:
            writer (ExcelWriter): Pandas Excel writer
            messages (list): List of processed messages
            target_user (str): Name of the target user
            my_name (str): Your name
        """
        # Prepare data for the conversation sheet
        conversation_data = []

        for msg in messages:
            # Media indicators
            media_info = []
            if msg['photos']:
                media_info.append(f"{len(msg['photos'])} photo(s)")
            if msg['videos']:
                media_info.append(f"{len(msg['videos'])} video(s)")
            if msg['audio']:
                media_info.append(f"{len(msg['audio'])} audio(s)")

            media_str = ", ".join(media_info) if media_info else ""

            # Reactions
            reactions_str = ", ".join(
                f"{r['reaction']} by {r['actor']}" for r in msg['reactions']
            ) if msg['reactions'] else ""

            # Fix any broken text in sender name and content
            sender = utils.fix_broken_text(msg['sender'])
            content = msg['content'] if msg['content'] else ""
            content = utils.fix_broken_text(content)

            # Add row to data
            conversation_data.append({
                'Date': msg['date'],
                'Time': msg['time'],
                'Sender': sender,
                'Message': content,
                'Media': media_str,
                'Reactions': reactions_str,
                'Has Emoji': "Yes" if msg['has_emoji'] else "No",
                'Emoji Count': msg['emoji_count'],
                'Is Good Morning': "Yes" if msg['is_good_morning'] else "No",
                'Mentions My Name': "Yes" if msg['mentions_my_name'] else "No",
                'Mentions Target Name': "Yes" if msg['mentions_target_name'] else "No",
                'Has Algerian Slang': "Yes" if msg['has_algerian_slang'] else "No"
            })

        # Create DataFrame
        df = pd.DataFrame(conversation_data)

        # Write to Excel
        df.to_excel(writer, sheet_name='Conversation', index=False)

    def _create_statistics_sheet(self, writer, stats, target_user, my_name):
        """
        Create the statistics sheet in the Excel file.

        Args:
            writer (ExcelWriter): Pandas Excel writer
            stats (dict): Statistics dictionary
            target_user (str): Name of the target user
            my_name (str): Your name
        """
        # Prepare data for the statistics sheet
        stats_data = [
            ['Conversation Statistics', ''],
            ['', ''],
            ['Total Messages', stats['total_messages']],
        ]

        # Add messages by sender
        for sender, count in stats['messages_by_sender'].items():
            stats_data.append([f'Messages from {sender}', count])

        # Add other statistics
        stats_data.extend([
            ['Total Emojis Used', stats['total_emojis']],
            ['Unique Emojis Count', stats['unique_emojis_count']],
            ['Good Morning Messages', stats['good_morning_count']],
            [f"Mentions of '{my_name}'", stats['my_name_mentions']],
            [f"Mentions of '{target_user}'", stats['target_name_mentions']],
            ['Active Conversation Days', stats['active_conversation_days']],
            ['First Message Date', stats['first_message_date']],
            ['Last Message Date', stats['last_message_date']],
            ['Conversation Duration (days)', stats['conversation_duration_days']],
            ['Algerian Slang Expressions', stats['algerian_slang_count']],
            ['Most Active Day', f"{stats['most_active_day']} ({stats['most_active_day_count']} messages)"],
            ['', ''],
            ['Top Emojis Used', ', '.join(stats['unique_emojis'][:20]) if stats['unique_emojis'] else 'None']
        ])

        # Create DataFrame
        df = pd.DataFrame(stats_data)

        # Write to Excel
        df.to_excel(writer, sheet_name='Statistics', header=False, index=False)

    def _create_media_sheet(self, writer, messages):
        """
        Create the media sheet in the Excel file.

        Args:
            writer (ExcelWriter): Pandas Excel writer
            messages (list): List of processed messages
        """
        # Prepare data for the media sheet
        media_data = []

        for msg in messages:
            # Fix any broken text in sender name
            sender = utils.fix_broken_text(msg['sender'])

            # Add photos
            for i, photo_uri in enumerate(msg['photos']):
                media_data.append({
                    'Date': msg['date'],
                    'Time': msg['time'],
                    'Sender': sender,
                    'Type': 'Photo',
                    'URI': photo_uri
                })

            # Add videos
            for i, video_uri in enumerate(msg['videos']):
                media_data.append({
                    'Date': msg['date'],
                    'Time': msg['time'],
                    'Sender': sender,
                    'Type': 'Video',
                    'URI': video_uri
                })

            # Add audio
            for i, audio_uri in enumerate(msg['audio']):
                media_data.append({
                    'Date': msg['date'],
                    'Time': msg['time'],
                    'Sender': sender,
                    'Type': 'Audio',
                    'URI': audio_uri
                })

        # Create DataFrame
        df = pd.DataFrame(media_data)

        # Write to Excel only if there's media data
        if not df.empty:
            df.to_excel(writer, sheet_name='Media Files', index=False)
        else:
            # Create an empty sheet with headers
            pd.DataFrame(columns=['Date', 'Time', 'Sender', 'Type', 'URI']).to_excel(
                writer, sheet_name='Media Files', index=False
            )

    def _apply_styling(self, writer):
        """
        Apply styling to the Excel workbook.

        Args:
            writer (ExcelWriter): Pandas Excel writer
        """
        workbook = writer.book

        # Define styles
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Apply styles to each sheet
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]

            # Style the header row
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[column_letter].width = min(adjusted_width, 50)
